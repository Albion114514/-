# -*- coding: utf-8 -*-
"""
Baidu Hot Extended - enhanced output version
会在当前目录下创建文件夹：
  Baidu_hot_extended_{YYYYMMDD_HHMMSS}
并将 Excel / CSV / JSON 文件全部保存其中。
"""

import os
import time
import json
import csv
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, Alignment


BAIDU_TOP_URL = "https://top.baidu.com/board?tab=realtime"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Cache-Control": "no-cache",
}


def build_session() -> requests.Session:
    s = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=0.6,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD", "OPTIONS"],
        raise_on_status=False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.headers.update(HEADERS)
    return s


def parse_items(soup: BeautifulSoup) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    cards = soup.select("div.category-wrap_iQLoo") or soup.select("div.category-wrap") or []
    now = datetime.now(timezone(timedelta(hours=8)))
    ts = now.strftime("%Y-%m-%d %H:%M:%S %z")

    for idx, card in enumerate(cards, start=1):
        title_el = card.select_one(".c-single-text-ellipsis")
        title = title_el.get_text(strip=True) if title_el else ""

        rank_el = card.select_one(".index_1Ew5p, .index")
        try:
            rank = int(rank_el.get_text(strip=True)) if rank_el else idx
        except Exception:
            rank = idx

        heat_el = card.select_one(".hot-index_1Bl1a, .hot-index")
        heat = heat_el.get_text(strip=True) if heat_el else ""

        desc_el = card.select_one(".hot-desc_1m_jR, .hot-desc")
        brief = desc_el.get_text(" ", strip=True) if desc_el else ""

        tag_el = card.select_one(".tag_1z8Gk, .tag")
        tag = tag_el.get_text(strip=True) if tag_el else ""

        link_el = card.select_one("a[href]")
        link = link_el["href"].strip() if link_el and link_el.has_attr("href") else ""

        trend_el = card.select_one(".trend-icon, .trend")
        trend = trend_el.get_text(strip=True) if trend_el else ""

        records.append({
            "rank": rank,
            "title": title,
            "heat": heat,
            "tag": tag,
            "brief": brief,
            "link": link,
            "trend": trend,
            "fetched_at": ts,
            "source": BAIDU_TOP_URL,
        })
    return records


def save_excel(rows: List[Dict[str, Any]], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baidu Hot Realtime"

    headers = ["rank", "title", "heat", "tag", "brief", "link", "trend", "fetched_at", "source"]
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = center

    col_widths = {"A": 6, "B": 42, "C": 10, "D": 12, "E": 66, "F": 50, "G": 8, "H": 20, "I": 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for i, cell in enumerate(row, start=1):
            cell.alignment = center if i in (1, 3, 7) else left

    wb.save(path)


def save_csv(rows: List[Dict[str, Any]], path: str) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def save_json(rows: List[Dict[str, Any]], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def main():
    # 创建带时间戳的输出文件夹
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = f"Baidu_hot_extended_{ts}"
    os.makedirs(folder, exist_ok=True)

    session = build_session()
    resp = session.get(BAIDU_TOP_URL, timeout=10)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    data = [d for d in parse_items(soup) if d.get("title")]

    excel_path = os.path.join(folder, f"Baidu_hot_extended_{ts}.xlsx")
    csv_path = os.path.join(folder, f"Baidu_hot_extended_{ts}.csv")
    json_path = os.path.join(folder, f"Baidu_hot_extended_{ts}.json")

    save_excel(data, excel_path)
    save_csv(data, csv_path)
    save_json(data, json_path)

    print(f"✅ 已保存 {len(data)} 条记录到文件夹：{folder}")
    print("文件列表：")
    print(f" - Excel：{excel_path}")
    print(f" - CSV：{csv_path}")
    print(f" - JSON：{json_path}")


if __name__ == "__main__":
    main()
